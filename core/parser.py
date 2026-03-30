from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


SYNTHETIC_HEADER_PREFIX = "column_"
EMPTY_LIKE_VALUES = {"", "nan", "none", "null", "n/a", "n\\a"}
TRIVIAL_NOTE_VALUES = {"yes", "no", "y", "n", "ok", "agree", "agreed"}
HANDLE_HINT_TOKENS = {"handle", "username", "user", "account", "id"}
URL_HINT_TOKENS = {"url", "link", "profile", "href"}
PLATFORM_METRIC_TOKENS = {
    "follower",
    "followers",
    "engagement",
    "impressions",
    "reach",
    "views",
    "likes",
    "comments",
    "shares",
    "avg",
    "average",
    "rate",
    "score",
    "count",
}

US_STATE_MAP: dict[str, str] = {
    "alabama": "AL",
    "alaska": "AK",
    "arizona": "AZ",
    "arkansas": "AR",
    "california": "CA",
    "colorado": "CO",
    "connecticut": "CT",
    "delaware": "DE",
    "district of columbia": "DC",
    "florida": "FL",
    "georgia": "GA",
    "hawaii": "HI",
    "idaho": "ID",
    "illinois": "IL",
    "indiana": "IN",
    "iowa": "IA",
    "kansas": "KS",
    "kentucky": "KY",
    "louisiana": "LA",
    "maine": "ME",
    "maryland": "MD",
    "massachusetts": "MA",
    "michigan": "MI",
    "minnesota": "MN",
    "mississippi": "MS",
    "missouri": "MO",
    "montana": "MT",
    "nebraska": "NE",
    "nevada": "NV",
    "new hampshire": "NH",
    "new jersey": "NJ",
    "new mexico": "NM",
    "new york": "NY",
    "north carolina": "NC",
    "north dakota": "ND",
    "ohio": "OH",
    "oklahoma": "OK",
    "oregon": "OR",
    "pennsylvania": "PA",
    "rhode island": "RI",
    "south carolina": "SC",
    "south dakota": "SD",
    "tennessee": "TN",
    "texas": "TX",
    "utah": "UT",
    "vermont": "VT",
    "virginia": "VA",
    "washington": "WA",
    "west virginia": "WV",
    "wisconsin": "WI",
    "wyoming": "WY",
}


@dataclass(slots=True)
class InfluencerRecord:
    """Normalized influencer row parsed from a Later export."""

    source_row_index: int
    canonical_name: str
    raw_name: str | None
    email: str | None
    phone_raw: str | None
    phone_digits: str | None
    phone_display: str | None
    city: str | None
    state_raw: str | None
    state: str | None
    location_display: str | None
    instagram_handle: str | None
    tiktok_handle: str | None
    youtube_handle: str | None
    facebook_handle: str | None
    instagram_url: str | None
    tiktok_url: str | None
    youtube_url: str | None
    facebook_url: str | None
    has_instagram: bool
    has_tiktok: bool
    has_youtube: bool
    has_facebook: bool
    platform_display_candidate: str | None
    primary_handle_candidate: str | None
    multi_handle_candidate: str | None
    raw_custom_responses: dict[str, str]
    combined_response_text: str | None
    notes_candidate_parts: list[str]
    raw_row: dict[str, object]
    warnings: list[str]


@dataclass(slots=True)
class ParseSummary:
    """Aggregate parser statistics and warnings."""

    source_filename: str
    worksheet_name: str
    total_rows_seen: int
    parsed_rows: int
    skipped_empty_rows: int
    skipped_invalid_rows: int
    duplicate_identity_candidates: int
    missing_name_rows: int
    missing_email_rows: int
    missing_handle_rows: int
    missing_location_rows: int
    record_warnings_count: int
    unknown_headers: list[str]
    global_warnings: list[str]


@dataclass(slots=True)
class ParseResult:
    """Top-level parse output consumed by downstream pipeline stages."""

    records: list[InfluencerRecord]
    summary: ParseSummary
    header_map: dict[str, str]
    worksheet_name: str


def parse_later_export(file_path: str | Path) -> ParseResult:
    """Parse a Later export workbook into normalized influencer records."""
    source_path = Path(file_path)
    if not source_path.exists() or not source_path.is_file():
        raise FileNotFoundError(f"Later export file not found: {source_path}")

    workbook: Workbook | None = None
    try:
        workbook = load_workbook(source_path, data_only=True)
    except Exception as exc:  # pragma: no cover
        raise ValueError(f"Workbook could not be opened: {source_path}") from exc

    try:
        worksheet = select_worksheet(workbook)
        header_row_index, raw_headers = detect_header_row(worksheet)
        header_map, unknown_headers = normalize_headers(raw_headers)

        records: list[InfluencerRecord] = []
        total_rows_seen = 0
        skipped_empty_rows = 0
        skipped_invalid_rows = 0

        for row_index, row_values in iter_data_rows(
            worksheet=worksheet,
            min_row=header_row_index + 1,
            max_col=len(raw_headers),
        ):
            total_rows_seen += 1
            row_data = row_to_raw_mapping(raw_headers, row_values)
            if detect_empty_row(row_data):
                skipped_empty_rows += 1
                continue

            record = parse_influencer_row(
                row_index=row_index,
                row_data=row_data,
                header_map=header_map,
            )
            if record is None:
                skipped_invalid_rows += 1
                continue
            records.append(record)

        if not records:
            raise ValueError(
                f"No usable data rows found in worksheet '{worksheet.title}'."
            )

        duplicate_identity_candidates = detect_duplicate_identity_candidates(records)
        global_warnings = build_global_warnings(
            unknown_headers=unknown_headers,
            skipped_invalid_rows=skipped_invalid_rows,
            duplicate_identity_candidates=duplicate_identity_candidates,
        )

        summary = build_parse_summary(
            source_filename=source_path.name,
            worksheet_name=worksheet.title,
            records=records,
            total_rows_seen=total_rows_seen,
            skipped_empty_rows=skipped_empty_rows,
            skipped_invalid_rows=skipped_invalid_rows,
            duplicate_identity_candidates=duplicate_identity_candidates,
            unknown_headers=unknown_headers,
            global_warnings=global_warnings,
        )

        return ParseResult(
            records=records,
            summary=summary,
            header_map=header_map,
            worksheet_name=worksheet.title,
        )
    finally:
        if workbook is not None:
            workbook.close()


def select_worksheet(workbook: Workbook) -> Worksheet:
    """Select the most likely usable worksheet from a workbook."""
    active_sheet = workbook.active
    ordered_sheets = [active_sheet] + [
        worksheet
        for worksheet in workbook.worksheets
        if worksheet.title != active_sheet.title
    ]

    any_header_found = False
    for worksheet in ordered_sheets:
        try:
            header_row_index, raw_headers = detect_header_row(worksheet)
        except ValueError:
            continue
        any_header_found = True
        if worksheet_has_data(worksheet, header_row_index, len(raw_headers)):
            return worksheet

    if not any_header_found:
        raise ValueError("No usable header row found in workbook.")
    raise ValueError("No usable worksheet found in workbook.")


def worksheet_is_usable(worksheet: Worksheet) -> bool:
    """Return whether a worksheet has a plausible header and at least one data row."""
    try:
        header_row_index, raw_headers = detect_header_row(worksheet)
    except ValueError:
        return False
    return worksheet_has_data(worksheet, header_row_index, len(raw_headers))


def detect_header_row(worksheet: Worksheet) -> tuple[int, list[str]]:
    """Detect the first plausible header row on a worksheet."""
    max_scan_rows = min(worksheet.max_row, 100)
    for row_index in range(1, max_scan_rows + 1):
        row_values = [
            worksheet.cell(row=row_index, column=column_index).value
            for column_index in range(1, worksheet.max_column + 1)
        ]
        cleaned_values = [normalize_string(value) for value in row_values]
        non_empty_values = [value for value in cleaned_values if value is not None]
        if len(non_empty_values) < 2:
            continue

        alpha_cell_count = sum(
            1 for value in non_empty_values if any(ch.isalpha() for ch in value)
        )
        if alpha_cell_count < 1:
            continue

        raw_headers = build_raw_headers(row_values)
        if len(raw_headers) < 2:
            continue
        return row_index, raw_headers

    raise ValueError(f"No usable header row found in worksheet '{worksheet.title}'.")


def normalize_headers(raw_headers: list[str]) -> tuple[dict[str, str], list[str]]:
    """Map raw source headers to stable internal parser keys."""
    header_map: dict[str, str] = {}
    unknown_headers: list[str] = []
    custom_index = 1

    for raw_header in raw_headers:
        normalized_key = map_header_to_internal_key(raw_header)
        if normalized_key is None:
            normalized_key = f"custom_response_{custom_index}"
            custom_index += 1
            if not raw_header.startswith(SYNTHETIC_HEADER_PREFIX):
                unknown_headers.append(raw_header)
        header_map[raw_header] = normalized_key

    return header_map, dedupe_preserve_order(unknown_headers)


def row_to_raw_mapping(raw_headers: list[str], row_values: list[object]) -> dict[str, object]:
    """Build a dictionary of raw header name to cell value for a row."""
    return {
        header: row_values[index] if index < len(row_values) else None
        for index, header in enumerate(raw_headers)
    }


def parse_influencer_row(
    row_index: int,
    row_data: dict[str, object],
    header_map: dict[str, str],
) -> InfluencerRecord | None:
    """Parse a raw data row into a normalized influencer record."""
    if detect_empty_row(row_data):
        return None

    name_source = first_mapped_value("full_name", row_data, header_map)
    raw_name = normalize_string(name_source)
    normalized_name = normalize_name(name_source)
    email = normalize_email(first_mapped_value("email", row_data, header_map))
    phone_raw, phone_digits, phone_display = normalize_phone(
        first_mapped_value("phone", row_data, header_map)
    )

    city = normalize_city(first_mapped_value("city", row_data, header_map))
    state_raw, state = normalize_state(first_mapped_value("state", row_data, header_map))
    location_display = build_location_display(city, state)

    instagram_handle = normalize_handle(
        first_mapped_value("instagram_handle", row_data, header_map)
    )
    tiktok_handle = normalize_handle(
        first_mapped_value("tiktok_handle", row_data, header_map)
    )
    youtube_handle = normalize_handle(
        first_mapped_value("youtube_handle", row_data, header_map)
    )
    facebook_handle = normalize_handle(
        first_mapped_value("facebook_handle", row_data, header_map)
    )

    instagram_url = normalize_url(first_mapped_value("instagram_url", row_data, header_map))
    tiktok_url = normalize_url(first_mapped_value("tiktok_url", row_data, header_map))
    youtube_url = normalize_url(first_mapped_value("youtube_url", row_data, header_map))
    facebook_url = normalize_url(first_mapped_value("facebook_url", row_data, header_map))

    has_instagram, has_tiktok, has_youtube, has_facebook = derive_platform_flags(
        instagram_handle=instagram_handle,
        tiktok_handle=tiktok_handle,
        youtube_handle=youtube_handle,
        facebook_handle=facebook_handle,
    )

    platform_display_candidate = build_platform_display_candidate(
        has_instagram=has_instagram,
        has_tiktok=has_tiktok,
        has_youtube=has_youtube,
        has_facebook=has_facebook,
    )

    primary_handle_candidate = build_primary_handle_candidate(
        instagram_handle=instagram_handle,
        tiktok_handle=tiktok_handle,
        youtube_handle=youtube_handle,
        facebook_handle=facebook_handle,
    )

    multi_handle_candidate = build_multi_handle_candidate(
        instagram_handle=instagram_handle,
        tiktok_handle=tiktok_handle,
        youtube_handle=youtube_handle,
        facebook_handle=facebook_handle,
    )

    raw_custom_responses = extract_custom_responses(row_data, header_map)
    combined_response_text = (
        "\n".join(raw_custom_responses.values()) if raw_custom_responses else None
    )
    notes_candidate_parts = extract_notes_candidate_parts(raw_custom_responses.values())

    has_any_handle = any([has_instagram, has_tiktok, has_youtube, has_facebook])
    has_meaningful_custom = bool(notes_candidate_parts)
    if (
        normalized_name is None
        and email is None
        and not has_any_handle
        and not has_meaningful_custom
    ):
        return None

    canonical_name = build_canonical_name(
        normalized_name=normalized_name,
        email=email,
        primary_handle_candidate=primary_handle_candidate,
        row_index=row_index,
    )

    warnings = collect_record_warnings(
        raw_name=raw_name,
        email=email,
        has_any_handle=has_any_handle,
        location_display=location_display,
        notes_candidate_parts=notes_candidate_parts,
    )

    return InfluencerRecord(
        source_row_index=row_index,
        canonical_name=canonical_name,
        raw_name=raw_name,
        email=email,
        phone_raw=phone_raw,
        phone_digits=phone_digits,
        phone_display=phone_display,
        city=city,
        state_raw=state_raw,
        state=state,
        location_display=location_display,
        instagram_handle=instagram_handle,
        tiktok_handle=tiktok_handle,
        youtube_handle=youtube_handle,
        facebook_handle=facebook_handle,
        instagram_url=instagram_url,
        tiktok_url=tiktok_url,
        youtube_url=youtube_url,
        facebook_url=facebook_url,
        has_instagram=has_instagram,
        has_tiktok=has_tiktok,
        has_youtube=has_youtube,
        has_facebook=has_facebook,
        platform_display_candidate=platform_display_candidate,
        primary_handle_candidate=primary_handle_candidate,
        multi_handle_candidate=multi_handle_candidate,
        raw_custom_responses=raw_custom_responses,
        combined_response_text=combined_response_text,
        notes_candidate_parts=notes_candidate_parts,
        raw_row=dict(row_data),
        warnings=warnings,
    )


def detect_empty_row(row_data: dict[str, object]) -> bool:
    """Return True when a row has no meaningful values across all columns."""
    return all(normalize_string(value) is None for value in row_data.values())


def normalize_string(value: object) -> str | None:
    """Normalize general string-like values and collapse empty placeholders."""
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    text = re.sub(r"\s+", " ", text)
    if text.lower() in EMPTY_LIKE_VALUES:
        return None
    return text


def normalize_name(value: object) -> str | None:
    """Normalize name values without forcing case transformation."""
    text = normalize_string(value)
    if text is None:
        return None
    if not any(ch.isalnum() for ch in text):
        return None
    return text


def normalize_email(value: object) -> str | None:
    """Normalize email values with basic structural sanity checks."""
    email = normalize_string(value)
    if email is None:
        return None

    email = email.lower()
    if email.count("@") != 1:
        return None
    if email.startswith("@") or email.endswith("@"):
        return None
    return email


def normalize_phone(value: object) -> tuple[str | None, str | None, str | None]:
    """Normalize phone values into raw, digits, and display candidates."""
    phone_raw = normalize_string(value)
    if phone_raw is None:
        return None, None, None

    phone_digits = re.sub(r"\D", "", phone_raw)
    if not phone_digits:
        return phone_raw, None, None

    if len(phone_digits) == 10:
        return phone_raw, phone_digits, format_phone_display(phone_digits)
    if len(phone_digits) == 11 and phone_digits.startswith("1"):
        return phone_raw, phone_digits, format_phone_display(phone_digits[-10:])
    return phone_raw, phone_digits, None


def normalize_handle(value: object) -> str | None:
    """Normalize social handle text to plain username without leading @."""
    handle = normalize_string(value)
    if handle is None:
        return None

    handle = re.sub(r"\s+", "", handle)
    handle = handle.lstrip("@")
    return handle or None


def normalize_url(value: object) -> str | None:
    """Normalize URL-like text while preserving source format."""
    return normalize_string(value)


def normalize_city(value: object) -> str | None:
    """Normalize city text values."""
    return normalize_string(value)


def normalize_state(value: object) -> tuple[str | None, str | None]:
    """Normalize US state text and return both raw and standardized values."""
    state_raw = normalize_string(value)
    if state_raw is None:
        return None, None

    if re.fullmatch(r"[A-Za-z]{2}", state_raw):
        return state_raw, state_raw.upper()

    normalized_key = re.sub(r"\s+", " ", state_raw).lower().replace(".", "")
    state = US_STATE_MAP.get(normalized_key)
    return state_raw, state


def build_location_display(city: str | None, state: str | None) -> str | None:
    """Build display location from normalized city/state values."""
    if city and state:
        return f"{city}, {state}"
    if city:
        return city
    if state:
        return state
    return None


def derive_platform_flags(
    instagram_handle: str | None,
    tiktok_handle: str | None,
    youtube_handle: str | None,
    facebook_handle: str | None,
) -> tuple[bool, bool, bool, bool]:
    """Derive platform presence booleans from normalized handles."""
    return (
        instagram_handle is not None,
        tiktok_handle is not None,
        youtube_handle is not None,
        facebook_handle is not None,
    )


def build_platform_display_candidate(
    has_instagram: bool,
    has_tiktok: bool,
    has_youtube: bool,
    has_facebook: bool,
) -> str | None:
    """Build compact platform display candidate using IG/TT/YT/FB ordering."""
    labels: list[str] = []
    if has_instagram:
        labels.append("IG")
    if has_tiktok:
        labels.append("TT")
    if has_youtube:
        labels.append("YT")
    if has_facebook:
        labels.append("FB")
    if not labels:
        return None
    return " + ".join(labels)


def build_primary_handle_candidate(
    instagram_handle: str | None,
    tiktok_handle: str | None,
    youtube_handle: str | None,
    facebook_handle: str | None,
) -> str | None:
    """Build prioritized single-handle candidate for downstream usage."""
    handle = instagram_handle or tiktok_handle or youtube_handle or facebook_handle
    return f"@{handle}" if handle else None


def build_multi_handle_candidate(
    instagram_handle: str | None,
    tiktok_handle: str | None,
    youtube_handle: str | None,
    facebook_handle: str | None,
) -> str | None:
    """Build single or multi-line handle display candidate."""
    entries: list[tuple[str, str]] = []
    if instagram_handle:
        entries.append(("IG", instagram_handle))
    if tiktok_handle:
        entries.append(("TT", tiktok_handle))
    if youtube_handle:
        entries.append(("YT", youtube_handle))
    if facebook_handle:
        entries.append(("FB", facebook_handle))

    if not entries:
        return None
    if len(entries) == 1:
        platform, handle = entries[0]
        if platform == "IG":
            return f"@{handle}"
        return f"{platform}- @{handle}"
    return "\n".join(f"{platform}- @{handle}" for platform, handle in entries)


def extract_custom_responses(
    row_data: dict[str, object],
    header_map: dict[str, str],
) -> dict[str, str]:
    """Extract and preserve non-empty custom response values from a row."""
    custom_responses: dict[str, str] = {}
    for raw_header, normalized_key in header_map.items():
        if not normalized_key.startswith("custom_response_"):
            continue
        cleaned = normalize_string(row_data.get(raw_header))
        if cleaned is not None:
            custom_responses[raw_header] = cleaned
    return custom_responses


def extract_notes_candidate_parts(values: Iterable[str]) -> list[str]:
    """Extract conservative, useful note fragments from custom responses."""
    parts: list[str] = []
    seen: set[str] = set()

    for value in values:
        fragments = re.split(r"[\r\n;]+", value)
        for fragment in fragments:
            cleaned = normalize_string(fragment)
            if cleaned is None:
                continue
            lowered = cleaned.lower()
            if lowered in TRIVIAL_NOTE_VALUES:
                continue
            if len(cleaned) < 2:
                continue
            if lowered in seen:
                continue
            seen.add(lowered)
            parts.append(cleaned)

    return parts


def collect_record_warnings(
    raw_name: str | None,
    email: str | None,
    has_any_handle: bool,
    location_display: str | None,
    notes_candidate_parts: list[str],
) -> list[str]:
    """Collect soft validation warnings for a parsed influencer record."""
    warnings: list[str] = []

    if raw_name is None:
        warnings.append("missing_name")
    if email is None:
        warnings.append("missing_email")
    if not has_any_handle:
        warnings.append("missing_all_handles")
    if location_display is None:
        warnings.append("missing_location")

    confidence_points = 0
    if raw_name is not None:
        confidence_points += 1
    if email is not None:
        confidence_points += 1
    if has_any_handle:
        confidence_points += 1
    if location_display is not None:
        confidence_points += 1
    if notes_candidate_parts:
        confidence_points += 1
    if confidence_points < 2:
        warnings.append("suspicious_partial_row")

    return warnings


def build_parse_summary(
    source_filename: str,
    worksheet_name: str,
    records: list[InfluencerRecord],
    total_rows_seen: int,
    skipped_empty_rows: int,
    skipped_invalid_rows: int,
    duplicate_identity_candidates: int,
    unknown_headers: list[str],
    global_warnings: list[str],
) -> ParseSummary:
    """Build parse summary object from parsed records and counters."""
    missing_name_rows = sum("missing_name" in record.warnings for record in records)
    missing_email_rows = sum("missing_email" in record.warnings for record in records)
    missing_handle_rows = sum(
        "missing_all_handles" in record.warnings for record in records
    )
    missing_location_rows = sum("missing_location" in record.warnings for record in records)
    record_warnings_count = sum(len(record.warnings) for record in records)

    return ParseSummary(
        source_filename=source_filename,
        worksheet_name=worksheet_name,
        total_rows_seen=total_rows_seen,
        parsed_rows=len(records),
        skipped_empty_rows=skipped_empty_rows,
        skipped_invalid_rows=skipped_invalid_rows,
        duplicate_identity_candidates=duplicate_identity_candidates,
        missing_name_rows=missing_name_rows,
        missing_email_rows=missing_email_rows,
        missing_handle_rows=missing_handle_rows,
        missing_location_rows=missing_location_rows,
        record_warnings_count=record_warnings_count,
        unknown_headers=unknown_headers,
        global_warnings=global_warnings,
    )


def map_header_to_internal_key(raw_header: str) -> str | None:
    """Map a raw header text value to a stable internal parser key."""
    cleaned = normalize_string(raw_header)
    if cleaned is None:
        return None

    lowered = cleaned.lower()
    tokens = set(re.findall(r"[a-z0-9]+", lowered))
    compact = re.sub(r"[^a-z0-9]+", "", lowered)

    exact_map = {
        "name": "full_name",
        "fullname": "full_name",
        "influencername": "full_name",
        "creatorname": "full_name",
        "email": "email",
        "emailaddress": "email",
        "contactemail": "email",
        "businessemail": "email",
        "phone": "phone",
        "phonenumber": "phone",
        "mobile": "phone",
        "mobilenumber": "phone",
        "cell": "phone",
        "cellphone": "phone",
        "city": "city",
        "state": "state",
        "province": "state",
        "region": "state",
        "instagramhandle": "instagram_handle",
        "ighandle": "instagram_handle",
        "instagramusername": "instagram_handle",
        "igusername": "instagram_handle",
        "instagramurl": "instagram_url",
        "igurl": "instagram_url",
        "tiktokhandle": "tiktok_handle",
        "tthandle": "tiktok_handle",
        "tiktokusername": "tiktok_handle",
        "ttusername": "tiktok_handle",
        "tiktokurl": "tiktok_url",
        "tturl": "tiktok_url",
        "youtubehandle": "youtube_handle",
        "ythandle": "youtube_handle",
        "youtubeusername": "youtube_handle",
        "ytusername": "youtube_handle",
        "youtubeurl": "youtube_url",
        "yturl": "youtube_url",
        "facebookhandle": "facebook_handle",
        "fbhandle": "facebook_handle",
        "facebookusername": "facebook_handle",
        "fbusername": "facebook_handle",
        "facebookurl": "facebook_url",
        "fburl": "facebook_url",
    }
    if compact in exact_map:
        return exact_map[compact]

    if "email" in tokens:
        return "email"
    if tokens.intersection({"phone", "mobile", "cell", "telephone"}):
        return "phone"
    if "city" in tokens:
        return "city"
    if tokens.intersection({"state", "province"}):
        return "state"
    if is_name_header(tokens):
        return "full_name"

    platform = detect_platform(tokens)
    if platform:
        if tokens.intersection(PLATFORM_METRIC_TOKENS):
            return None
        if tokens.intersection(URL_HINT_TOKENS) or "url" in compact or "link" in compact:
            return f"{platform}_url"
        if tokens.intersection(HANDLE_HINT_TOKENS):
            return f"{platform}_handle"
        if len(tokens) <= 2:
            return f"{platform}_handle"

    return None


def is_name_header(tokens: set[str]) -> bool:
    """Return True when tokens represent a likely person-name column."""
    if tokens == {"name"}:
        return True
    if "name" in tokens and tokens.intersection({"full", "influencer", "creator", "profile"}):
        return True
    return False


def detect_platform(tokens: set[str]) -> str | None:
    """Detect social platform key from header tokens."""
    if tokens.intersection({"instagram", "ig", "insta"}):
        return "instagram"
    if tokens.intersection({"tiktok", "tt"}):
        return "tiktok"
    if tokens.intersection({"youtube", "yt"}):
        return "youtube"
    if tokens.intersection({"facebook", "fb"}):
        return "facebook"
    return None


def build_raw_headers(row_values: list[object]) -> list[str]:
    """Build unique raw header labels from header-row cell values."""
    headers = [
        normalize_string(value) or f"{SYNTHETIC_HEADER_PREFIX}{index}"
        for index, value in enumerate(row_values, start=1)
    ]
    return make_unique_headers(headers)


def make_unique_headers(headers: list[str]) -> list[str]:
    """Ensure header labels are unique while preserving source order."""
    result: list[str] = []
    seen: dict[str, int] = {}
    for header in headers:
        count = seen.get(header, 0) + 1
        seen[header] = count
        if count == 1:
            result.append(header)
        else:
            result.append(f"{header}__{count}")
    return result


def worksheet_has_data(
    worksheet: Worksheet,
    header_row_index: int,
    max_col: int,
) -> bool:
    """Return True when worksheet has at least one non-empty row after header."""
    for _, row_values in iter_data_rows(
        worksheet=worksheet,
        min_row=header_row_index + 1,
        max_col=max_col,
    ):
        if any(normalize_string(value) is not None for value in row_values):
            return True
    return False


def iter_data_rows(
    worksheet: Worksheet,
    min_row: int,
    max_col: int,
) -> Iterable[tuple[int, list[object]]]:
    """Yield worksheet row index and row values for table rows."""
    for row_index, row in enumerate(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ),
        start=min_row,
    ):
        yield row_index, list(row)


def first_mapped_value(
    key: str,
    row_data: dict[str, object],
    header_map: dict[str, str],
) -> object:
    """Return first non-empty raw value from headers mapped to the given key."""
    for raw_header, normalized_key in header_map.items():
        if normalized_key != key:
            continue
        value = row_data.get(raw_header)
        if normalize_string(value) is not None:
            return value
    return None


def build_canonical_name(
    normalized_name: str | None,
    email: str | None,
    primary_handle_candidate: str | None,
    row_index: int,
) -> str:
    """Build canonical row identity with deterministic fallback behavior."""
    if normalized_name:
        return normalized_name
    if primary_handle_candidate:
        return primary_handle_candidate
    if email:
        return email.split("@", maxsplit=1)[0]
    return f"Unknown {row_index}"


def detect_duplicate_identity_candidates(records: list[InfluencerRecord]) -> int:
    """Count approximate duplicate identity tuples across parsed records."""
    seen: set[tuple[str, str, str]] = set()
    duplicates = 0

    for record in records:
        name_key = record.canonical_name.strip().lower() if record.canonical_name else ""
        email_key = record.email.strip().lower() if record.email else ""
        handle_key = (
            record.primary_handle_candidate.strip().lower()
            if record.primary_handle_candidate
            else ""
        )
        parts_count = sum(bool(part) for part in (name_key, email_key, handle_key))
        if parts_count < 2:
            continue
        identity_key = (name_key, email_key, handle_key)
        if identity_key in seen:
            duplicates += 1
        else:
            seen.add(identity_key)

    return duplicates


def build_global_warnings(
    unknown_headers: list[str],
    skipped_invalid_rows: int,
    duplicate_identity_candidates: int,
) -> list[str]:
    """Build global parse warnings from summary-level conditions."""
    warnings: list[str] = []
    if unknown_headers:
        warnings.append(
            f"{len(unknown_headers)} header(s) were unmapped and preserved as custom responses."
        )
    if skipped_invalid_rows:
        warnings.append(f"{skipped_invalid_rows} row(s) were skipped as unusable.")
    if duplicate_identity_candidates:
        warnings.append(
            f"{duplicate_identity_candidates} potential duplicate identity row(s) detected."
        )
    return warnings


def format_phone_display(digits: str) -> str:
    """Format 10 digits into U.S. phone display format."""
    return f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"


def dedupe_preserve_order(values: list[str]) -> list[str]:
    """Deduplicate list values while preserving first-seen order."""
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result
