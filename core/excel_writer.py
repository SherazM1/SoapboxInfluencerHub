from __future__ import annotations

from dataclasses import asdict, dataclass, is_dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.mapper import MappingResult


@dataclass(slots=True)
class SectionLayout:
    """Writable location metadata for one worksheet section."""

    sheet_name: str
    start_row: int
    max_rows: int
    column_map: dict[str, int]


@dataclass(slots=True)
class WorkbookLayout:
    """Workbook-wide writable section definitions."""

    workflow_influencer_details: SectionLayout
    workflow_draft_stages: SectionLayout
    workflow_live_content: SectionLayout
    workflow_content_checks: SectionLayout
    round_1: SectionLayout
    recruiting: SectionLayout


@dataclass(slots=True)
class WriteResult:
    """Result of writing mapped data to an output workbook."""

    output_path: str
    sections_written: dict[str, int]


def get_default_workbook_layout() -> WorkbookLayout:
    """Return the default template layout for phase-1 writing."""
    return WorkbookLayout(
        workflow_influencer_details=SectionLayout(
            sheet_name="Workflow",
            start_row=3,
            max_rows=6,
            column_map={
                "influencer_name": 1,
                "handle_linked": 2,
                "phone_number": 3,
                "email": 4,
                "contract_sent": 5,
                "contract_signed": 6,
                "payment_amount": 7,
                "deliverable": 8,
                "product": 9,
                "invoiced": 10,
                "w9": 11,
                "notes": 12,
            },
        ),
        workflow_draft_stages=SectionLayout(
            sheet_name="Workflow",
            start_row=12,
            max_rows=20,
            column_map={
                "influencer_name": 1,
                "draft_due_date": 2,
                "draft_status": 3,
                "final_approval": 4,
                "notes": 5,
            },
        ),
        workflow_live_content=SectionLayout(
            sheet_name="Workflow",
            start_row=41,
            max_rows=21,
            column_map={
                "influencer_name": 1,
                "deliverable": 2,
                "live_date": 3,
                "live_date_confirmed": 4,
                "link_to_live_content": 5,
                "notes": 6,
            },
        ),
        workflow_content_checks=SectionLayout(
            sheet_name="Workflow",
            start_row=66,
            max_rows=21,
            column_map={
                "influencer_name": 1,
                "likes_unhidden": 2,
                "link_in_caption": 3,
                "link_in_bio": 4,
                "hashtag_lindt_love": 5,
                "hashtag_walmart": 6,
                "mention_lindt_usa": 7,
                "story_count": 10,
                "notes": 11,
            },
        ),
        round_1=SectionLayout(
            sheet_name="Influencer Rounds",
            start_row=3,
            max_rows=129,
            column_map={
                "influencer_name": 1,
                "handle": 2,
                "deliverable": 3,
                "location": 4,
                "product": 5,
                "notes": 6,
            },
        ),
        recruiting=SectionLayout(
            sheet_name="Influencer Rounds",
            start_row=3,
            max_rows=129,
            column_map={
                "influencer_name": 8,
                "handle": 9,
                "platform": 10,
                "email": 11,
                "ro": 12,
                "fo": 13,
                "notes": 14,
            },
        ),
    )


def write_campaign_workbook(
    template_path: str | Path,
    mapping_result: MappingResult,
    output_path: str | Path,
    layout: WorkbookLayout | None = None,
) -> WriteResult:
    """Write mapped campaign sections into a template workbook and save output."""
    template = Path(template_path)
    output = Path(output_path)
    workbook_layout = layout or get_default_workbook_layout()

    workbook = load_template_workbook(template)
    try:
        workflow_written = write_workflow_sections(
            workbook=workbook,
            mapping_result=mapping_result,
            layout=workbook_layout,
        )
        rounds_written = write_influencer_rounds(
            workbook=workbook,
            mapping_result=mapping_result,
            layout=workbook_layout,
        )

        ensure_output_directory(output)
        save_workbook(workbook=workbook, output_path=output)
    finally:
        workbook.close()

    sections_written = {**workflow_written, **rounds_written}
    return WriteResult(output_path=str(output), sections_written=sections_written)


def load_template_workbook(template_path: Path) -> Workbook:
    """Load template workbook or raise a clear runtime error."""
    if not template_path.exists() or not template_path.is_file():
        raise FileNotFoundError(f"Template workbook does not exist: {template_path}")
    try:
        return load_workbook(template_path)
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(f"Unable to open template workbook: {template_path}") from exc


def ensure_output_directory(output_path: Path) -> None:
    """Ensure output directory exists for target workbook path."""
    output_dir = output_path.parent
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            f"Unable to create output directory: {output_dir}"
        ) from exc


def save_workbook(workbook: Workbook, output_path: Path) -> None:
    """Save workbook to output path with clear error reporting."""
    try:
        workbook.save(output_path)
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(f"Unable to save workbook to: {output_path}") from exc


def write_workflow_sections(
    workbook: Workbook,
    mapping_result: MappingResult,
    layout: WorkbookLayout,
) -> dict[str, int]:
    """Write all workflow sections and return counts by section name."""
    counts: dict[str, int] = {}

    counts["workflow_influencer_details"] = write_section_rows(
        worksheet=get_worksheet_or_raise(
            workbook=workbook,
            sheet_name=layout.workflow_influencer_details.sheet_name,
        ),
        section_layout=layout.workflow_influencer_details,
        rows=mapping_result.workflow_influencer_details_rows,
    )
    counts["workflow_draft_stages"] = write_section_rows(
        worksheet=get_worksheet_or_raise(
            workbook=workbook,
            sheet_name=layout.workflow_draft_stages.sheet_name,
        ),
        section_layout=layout.workflow_draft_stages,
        rows=mapping_result.workflow_draft_stage_rows,
    )
    counts["workflow_live_content"] = write_section_rows(
        worksheet=get_worksheet_or_raise(
            workbook=workbook,
            sheet_name=layout.workflow_live_content.sheet_name,
        ),
        section_layout=layout.workflow_live_content,
        rows=mapping_result.workflow_live_content_rows,
    )
    counts["workflow_content_checks"] = write_section_rows(
        worksheet=get_worksheet_or_raise(
            workbook=workbook,
            sheet_name=layout.workflow_content_checks.sheet_name,
        ),
        section_layout=layout.workflow_content_checks,
        rows=mapping_result.workflow_content_check_rows,
    )

    return counts


def write_influencer_rounds(
    workbook: Workbook,
    mapping_result: MappingResult,
    layout: WorkbookLayout,
) -> dict[str, int]:
    """Write all influencer-round sections and return counts by section name."""
    counts: dict[str, int] = {}

    counts["round_1"] = write_section_rows(
        worksheet=get_worksheet_or_raise(
            workbook=workbook,
            sheet_name=layout.round_1.sheet_name,
        ),
        section_layout=layout.round_1,
        rows=mapping_result.round_1_rows,
    )
    counts["recruiting"] = write_section_rows(
        worksheet=get_worksheet_or_raise(
            workbook=workbook,
            sheet_name=layout.recruiting.sheet_name,
        ),
        section_layout=layout.recruiting,
        rows=mapping_result.recruiting_rows,
    )

    return counts


def write_section_rows(
    worksheet: Worksheet,
    section_layout: SectionLayout,
    rows: list[object],
) -> int:
    """Write one section's rows and clear unused writable cells."""
    validate_section_capacity(section_layout=section_layout, row_count=len(rows))

    for row_offset, row_object in enumerate(rows):
        write_row_to_sheet(
            worksheet=worksheet,
            row_number=section_layout.start_row + row_offset,
            row_object=row_object,
            column_map=section_layout.column_map,
        )

    clear_unused_section_rows(
        worksheet=worksheet,
        section_layout=section_layout,
        used_count=len(rows),
    )
    return len(rows)


def write_row_to_sheet(
    worksheet: Worksheet,
    row_number: int,
    row_object: object,
    column_map: dict[str, int],
) -> None:
    """Write one mapped row object into worksheet columns defined by column_map."""
    row_values = row_object_to_mapping(row_object)
    for field_name, column_index in column_map.items():
        value = row_values.get(field_name, "")
        worksheet.cell(row=row_number, column=column_index, value=value_or_blank(value))


def clear_unused_section_rows(
    worksheet: Worksheet,
    section_layout: SectionLayout,
    used_count: int,
) -> None:
    """Clear values for unused rows in this section's owned writable columns."""
    if used_count >= section_layout.max_rows:
        return

    start_row = section_layout.start_row + used_count
    end_row = section_layout.start_row + section_layout.max_rows - 1
    columns = sorted(set(section_layout.column_map.values()))

    for row_number in range(start_row, end_row + 1):
        for column_index in columns:
            worksheet.cell(row=row_number, column=column_index, value=None)


def get_worksheet_or_raise(workbook: Workbook, sheet_name: str) -> Worksheet:
    """Get worksheet by name or raise a clear error."""
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f'Expected worksheet "{sheet_name}" was not found in template.')
    return workbook[sheet_name]


def validate_section_capacity(section_layout: SectionLayout, row_count: int) -> None:
    """Validate that row count fits within section capacity."""
    if row_count > section_layout.max_rows:
        raise ValueError(
            f'Section overflow for "{section_layout.sheet_name}" starting at row '
            f"{section_layout.start_row}: row_count={row_count}, "
            f"max_rows={section_layout.max_rows}."
        )


def row_object_to_mapping(row_object: object) -> dict[str, Any]:
    """Convert a row object into a field/value mapping."""
    if isinstance(row_object, dict):
        return dict(row_object)
    if is_dataclass(row_object):
        return asdict(row_object)
    return {
        name: value
        for name, value in vars(row_object).items()
        if not name.startswith("_")
    }


def value_or_blank(value: Any) -> Any:
    """Return empty string for None, otherwise the provided value."""
    return "" if value is None else value
