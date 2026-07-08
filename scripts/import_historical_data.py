from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from core.db import get_db_connection, maybe_init_database
from core.historical_data import build_seed_campaign_payload

SHEET_NAME = "Data "


@dataclass
class ImportSummary:
    rows_read: int = 0
    rows_inserted: int = 0
    rows_skipped: int = 0
    rows_failed: int = 0


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def row_is_blank(values: tuple[Any, ...]) -> bool:
    return all(is_blank(value) for value in values)


def parse_campaign_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except (TypeError, ValueError, OverflowError):
            return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(cleaned).date()
        except ValueError:
            return None
    return None


def excel_row_to_payload(values: tuple[Any, ...]) -> dict[str, Any]:
    row = {
        "Program": values[0],
        "Date": values[1],
        "# of Influencers": values[2],
        "Engagements": values[3],
        "Organic Impressions": values[4],
        "Paid Impressions": values[7],
        "Paid Spend (Impressions)": values[8],
        "Paid Engagement": values[10],
        "Paid Spend (Engagement)": values[11],
        "Paid Clicks": values[13],
        "Paid Spend (Clicks)": values[14],
    }
    payload = build_seed_campaign_payload(row)
    payload["campaign_date"] = parse_campaign_date(payload["campaign_date"])
    return payload


def validate_payload(payload: dict[str, Any]) -> list[str]:
    errors = []
    if not payload.get("program_name"):
        errors.append("program_name is required")
    if payload.get("campaign_date") is None:
        errors.append("campaign_date is required")
    for key in ("influencer_count", "engagements", "organic_impressions"):
        if payload.get(key) is None:
            errors.append(f"{key} is required")
    return errors


def iter_excel_payloads(workbook_path: Path) -> tuple[list[dict[str, Any]], int]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f'Workbook must contain a sheet named "{SHEET_NAME}".')

    sheet = workbook[SHEET_NAME]
    payloads = []
    blank_rows = 0
    for values in sheet.iter_rows(min_row=2, max_col=16, values_only=True):
        if row_is_blank(values):
            blank_rows += 1
            continue
        payloads.append(excel_row_to_payload(values))
    workbook.close()
    return payloads, blank_rows


def existing_campaign_id(cursor: Any, payload: dict[str, Any]) -> Any:
    cursor.execute(
        """
        select id
        from campaigns
        where program_name = %s and campaign_date = %s
        """,
        (payload["program_name"], payload["campaign_date"]),
    )
    row = cursor.fetchone()
    return row["id"] if row else None


def campaign_has_metrics(cursor: Any, campaign_id: Any) -> bool:
    cursor.execute(
        "select 1 from campaign_metrics where campaign_id = %s",
        (campaign_id,),
    )
    return cursor.fetchone() is not None


def insert_campaign(cursor: Any, payload: dict[str, Any]) -> Any:
    cursor.execute(
        """
        insert into campaigns (program_name, campaign_date)
        values (%s, %s)
        on conflict (program_name, campaign_date) do nothing
        returning id
        """,
        (payload["program_name"], payload["campaign_date"]),
    )
    row = cursor.fetchone()
    if row:
        return row["id"]
    return existing_campaign_id(cursor, payload)


def insert_campaign_metrics(cursor: Any, campaign_id: Any, payload: dict[str, Any]) -> bool:
    cursor.execute(
        """
        insert into campaign_metrics (
            campaign_id,
            influencer_count,
            engagements,
            organic_impressions,
            paid_impressions,
            paid_spend_impressions,
            paid_engagements,
            paid_spend_engagements,
            paid_clicks,
            paid_spend_clicks
        )
        values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        on conflict (campaign_id) do nothing
        returning id
        """,
        (
            campaign_id,
            payload["influencer_count"],
            payload["engagements"],
            payload["organic_impressions"],
            payload["paid_impressions"],
            payload["paid_spend_impressions"],
            payload["paid_engagements"],
            payload["paid_spend_engagements"],
            payload["paid_clicks"],
            payload["paid_spend_clicks"],
        ),
    )
    return cursor.fetchone() is not None


def import_payloads(
    payloads: list[dict[str, Any]], blank_rows: int, dry_run: bool
) -> ImportSummary:
    summary = ImportSummary(rows_read=len(payloads), rows_skipped=blank_rows)
    if not dry_run and not maybe_init_database():
        raise RuntimeError("Database is not configured or schema initialization failed.")

    with get_db_connection() as connection:
        if connection is None:
            raise RuntimeError("Database connection is not available.")

        try:
            with connection.transaction():
                with connection.cursor() as cursor:
                    for payload in payloads:
                        errors = validate_payload(payload)
                        if errors:
                            summary.rows_failed += 1
                            continue

                        campaign_id = existing_campaign_id(cursor, payload)
                        if campaign_id is not None and campaign_has_metrics(
                            cursor, campaign_id
                        ):
                            summary.rows_skipped += 1
                            continue

                        if dry_run:
                            summary.rows_inserted += 1
                            continue

                        campaign_id = insert_campaign(cursor, payload)
                        if campaign_id is None:
                            summary.rows_failed += 1
                            continue

                        if insert_campaign_metrics(cursor, campaign_id, payload):
                            summary.rows_inserted += 1
                        else:
                            summary.rows_skipped += 1
        except Exception:
            connection.rollback()
            raise

    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import historical campaign rows from the Excel Data sheet."
    )
    parser.add_argument("workbook_path", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    payloads, blank_rows = iter_excel_payloads(args.workbook_path)
    summary = import_payloads(payloads, blank_rows, args.dry_run)
    print(f"rows read: {summary.rows_read}")
    print(f"rows inserted: {summary.rows_inserted}")
    print(f"rows skipped: {summary.rows_skipped}")
    print(f"rows failed: {summary.rows_failed}")


if __name__ == "__main__":
    main()
